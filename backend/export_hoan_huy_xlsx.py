from __future__ import annotations

import json
import sys
from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill


if hasattr(sys.stdin, "reconfigure"):
    sys.stdin.reconfigure(encoding="utf-8")
if hasattr(sys.stdout, "reconfigure"):
    sys.stdout.reconfigure(encoding="utf-8")
if hasattr(sys.stderr, "reconfigure"):
    sys.stderr.reconfigure(encoding="utf-8")


HEADERS = [
    "Shop",
    "Sàn",
    "Loại đơn",
    "Kho",
    "Xử lý",
    "Ưu tiên",
    "Mã vận đơn",
    "Ngày sàn",
    "Tuổi case",
    "Chi tiết",
    "Ghi chú",
]


def main() -> None:
    payload = json.load(sys.stdin)
    rows = payload.get("rows", [])

    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Don Hoan Huy"
    worksheet.append(HEADERS)

    header_fill = PatternFill(fill_type="solid", fgColor="D9F99D")
    header_font = Font(bold=True)

    for cell in worksheet[1]:
      cell.fill = header_fill
      cell.font = header_font

    for row in rows:
        worksheet.append(
            [
                row.get("shop", ""),
                "TikTok" if row.get("san") == "tiktok" else "Shopee",
                "Đơn hủy" if row.get("loaiDon") == "don_huy" else "Đơn hoàn",
                row.get("warehouseStatusLabel", ""),
                row.get("processingStatusLabel", ""),
                row.get("priorityLabel", ""),
                row.get("trackingCode", ""),
                row.get("ngaySanDisplay", ""),
                f"{row.get('ageDays', 0)} ngày",
                row.get("detailText", ""),
                row.get("note", ""),
            ]
        )

    widths = {
        "A": 22,
        "B": 12,
        "C": 14,
        "D": 16,
        "E": 16,
        "F": 12,
        "G": 24,
        "H": 20,
        "I": 12,
        "J": 40,
        "K": 24,
    }

    for column, width in widths.items():
        worksheet.column_dimensions[column].width = width

    output = BytesIO()
    workbook.save(output)
    sys.stdout.buffer.write(output.getvalue())


if __name__ == "__main__":
    main()
