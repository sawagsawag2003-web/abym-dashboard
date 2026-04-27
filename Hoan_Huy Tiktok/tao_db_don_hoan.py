import os
import re
import sqlite3
import sys
from pathlib import Path

from openpyxl import load_workbook


BASE_DIR = Path(__file__).resolve().parent
PROJECT_ROOT = BASE_DIR.parent
SOURCE_FOLDER = BASE_DIR
DB_PATH = PROJECT_ROOT / "backend" / "DonHoan_Tiktok.db"

if hasattr(sys.stdout, "reconfigure"):
    sys.stdout.reconfigure(encoding="utf-8")
if hasattr(sys.stderr, "reconfigure"):
    sys.stderr.reconfigure(encoding="utf-8")


def init_db(db_path: str):
    os.makedirs(os.path.dirname(db_path), exist_ok=True)

    conn = sqlite3.connect(db_path)
    cursor = conn.cursor()

    cursor.execute("""
        CREATE TABLE IF NOT EXISTS don_hoan (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            shop TEXT,
            loai_don TEXT,
            return_order_id TEXT,
            thoi_gian_tao TEXT,
            order_id TEXT,
            ma_van_don TEXT,
            phan_loai TEXT,
            ly_do_hoan TEXT,
            so_tien REAL,
            ten_khach TEXT,
            source_file TEXT,
            UNIQUE(shop, order_id, ma_van_don, source_file)
        )
    """)

    columns = {
        row[1]
        for row in cursor.execute("PRAGMA table_info(don_hoan)")
    }

    if "return_order_id" not in columns:
        cursor.execute("ALTER TABLE don_hoan ADD COLUMN return_order_id TEXT")
    if "loai_don" not in columns:
        cursor.execute("ALTER TABLE don_hoan ADD COLUMN loai_don TEXT")

    cursor.execute("""
        UPDATE don_hoan
        SET loai_don = CASE
            WHEN source_file LIKE '%Đã hủy đơn hàng%' THEN 'don_huy'
            WHEN source_file LIKE '%Đơn trả hàng_hoàn tiền%' THEN 'don_hoan'
            ELSE 'don_hoan'
        END
        WHERE loai_don IS NULL OR loai_don = ''
    """)

    cursor.execute("""
        DELETE FROM don_hoan
        WHERE id NOT IN (
            SELECT MIN(id)
            FROM don_hoan
            WHERE order_id IS NOT NULL
            GROUP BY shop, loai_don, order_id
        )
        AND order_id IS NOT NULL
    """)

    cursor.execute("DROP INDEX IF EXISTS idx_don_hoan_unique_order")
    cursor.execute("""
        CREATE UNIQUE INDEX IF NOT EXISTS idx_don_hoan_unique_order_type
        ON don_hoan(shop, loai_don, order_id)
        WHERE order_id IS NOT NULL
    """)

    conn.commit()
    return conn


def parse_money(value):
    if value is None or value == "":
        return None

    if isinstance(value, (int, float)):
        return float(value)

    text = str(value).strip()
    text = text.replace(",", "")
    text = text.replace(".", "")
    text = text.replace("₫", "")
    text = text.replace("VND", "")
    text = text.replace(" ", "")
    text = re.sub(r"[^0-9\-]", "", text)

    if text == "":
        return None

    try:
        return float(text)
    except ValueError:
        return None


def extract_shop_name(filename: str):
    base = Path(filename).stem
    return base.split("_")[0].strip()


def is_return_file(filename: str):
    return filename.lower().endswith(".xlsx") and "đơn trả hàng_hoàn tiền" in filename.lower()


def is_cancel_file(filename: str):
    return filename.lower().endswith(".xlsx") and "đã hủy đơn hàng" in filename.lower()


def read_return_file(file_path: str):
    wb = load_workbook(file_path, data_only=True, read_only=False)
    ws = wb.active

    rows = []
    shop = extract_shop_name(os.path.basename(file_path))

    for row in ws.iter_rows(min_row=2, values_only=True):
        return_order_id = row[0] if len(row) > 0 else None     # A
        order_id = row[1] if len(row) > 1 else None            # B
        phan_loai = row[9] if len(row) > 9 else None           # J
        ten_khach = row[10] if len(row) > 10 else None         # K
        thoi_gian_tao = row[12] if len(row) > 12 else None     # M
        ly_do_hoan = row[13] if len(row) > 13 else None        # N
        so_tien = row[14] if len(row) > 14 else None           # O
        ma_van_don = row[16] if len(row) > 16 else None        # Q

        if not order_id:
            continue

        rows.append((
            shop,
            "don_hoan",
            str(return_order_id) if return_order_id is not None else None,
            str(thoi_gian_tao) if thoi_gian_tao is not None else None,
            str(order_id) if order_id is not None else None,
            str(ma_van_don) if ma_van_don is not None else None,
            str(phan_loai) if phan_loai is not None else None,
            str(ly_do_hoan) if ly_do_hoan is not None else None,
            parse_money(so_tien),
            str(ten_khach) if ten_khach is not None else None,
            os.path.basename(file_path),
        ))

    wb.close()
    return rows


def read_cancel_file(file_path: str):
    wb = load_workbook(file_path, data_only=True, read_only=False)
    ws = wb.active

    rows = []
    shop = extract_shop_name(os.path.basename(file_path))

    for row in ws.iter_rows(min_row=3, values_only=True):
        order_id = row[0] if len(row) > 0 else None            # A
        phan_loai = row[8] if len(row) > 8 else None           # I
        so_tien = row[15] if len(row) > 15 else None           # P
        thoi_gian_tao = row[24] if len(row) > 24 else None     # Y
        ly_do_hoan = row[31] if len(row) > 31 else None        # AF
        ma_van_don = row[34] if len(row) > 34 else None        # AI
        ten_khach = row[38] if len(row) > 38 else None         # AM

        if not order_id or not ma_van_don:
            continue

        rows.append((
            shop,
            "don_huy",
            None,
            str(thoi_gian_tao) if thoi_gian_tao is not None else None,
            str(order_id) if order_id is not None else None,
            str(ma_van_don) if ma_van_don is not None else None,
            str(phan_loai) if phan_loai is not None else None,
            str(ly_do_hoan) if ly_do_hoan is not None else None,
            parse_money(so_tien),
            str(ten_khach) if ten_khach is not None else None,
            os.path.basename(file_path),
        ))

    wb.close()
    return rows


def insert_rows(conn, data):
    cursor = conn.cursor()
    inserted = 0
    updated = 0

    for row in data:
        shop = row[0]
        loai_don = row[1]
        order_id = row[4]

        cursor.execute("""
            UPDATE don_hoan
            SET return_order_id = ?,
                thoi_gian_tao = ?,
                ma_van_don = ?,
                phan_loai = ?,
                ly_do_hoan = ?,
                so_tien = ?,
                ten_khach = ?,
                source_file = ?
            WHERE shop = ? AND loai_don = ? AND order_id = ?
        """, (
            row[2],
            row[3],
            row[5],
            row[6],
            row[7],
            row[8],
            row[9],
            row[10],
            shop,
            loai_don,
            order_id,
        ))

        if cursor.rowcount:
            updated += cursor.rowcount
            continue

        cursor.execute("""
            INSERT INTO don_hoan (
                shop,
                loai_don,
                return_order_id,
                thoi_gian_tao,
                order_id,
                ma_van_don,
                phan_loai,
                ly_do_hoan,
                so_tien,
                ten_khach,
                source_file
            )
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        """, row)
        inserted += 1

    conn.commit()
    return inserted, updated


def main():
    if not os.path.exists(SOURCE_FOLDER):
        print(f"Không tìm thấy thư mục nguồn: {SOURCE_FOLDER}")
        return

    conn = init_db(DB_PATH)

    total_files = 0
    total_rows = 0

    for filename in os.listdir(SOURCE_FOLDER):
        file_path = os.path.join(SOURCE_FOLDER, filename)

        try:
            if is_return_file(filename):
                rows = read_return_file(file_path)
            elif is_cancel_file(filename):
                rows = read_cancel_file(file_path)
            else:
                continue

            inserted, updated = insert_rows(conn, rows)

            total_files += 1
            total_rows += len(rows)

            print(f"Đã xử lý: {filename} | đọc {len(rows)} dòng | thêm {inserted} | cập nhật {updated}")
        except Exception as e:
            print(f"Lỗi file {filename}: {e}")

    conn.close()

    print("-" * 50)
    print("Hoàn tất.")
    print(f"Số file đã xử lý: {total_files}")
    print(f"Tổng số dòng đã đọc: {total_rows}")
    print(f"Database lưu tại: {DB_PATH}")


if __name__ == "__main__":
    main()
